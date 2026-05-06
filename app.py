import streamlit as st
import pandas as pd
import re
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
        page_title="SA Report Automation V7",
        layout="wide"
    )

st.title("📊 SA Report Automation V7")
# =====================================================
# PHASE 1 — FILE UPLOAD + CLEANING + READY FOR NEXT STEP
# =====================================================
def run_sa_report():



    # =====================================================
    # UPLOAD FILES
    # =====================================================
    uploaded_files = st.file_uploader(
        "📂 Upload All Files",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True
    )

    if not uploaded_files:
        st.info("Please upload files first.")
        st.stop()

    file_names = [f.name for f in uploaded_files]

    # =====================================================
    # FILE SELECTION
    # =====================================================
    st.subheader("📝 Select Required Files")

    col1, col2, col3 = st.columns(3)

    with col1:
        media_plan_name = st.selectbox(
            "📌 Media Plan",
            ["-- Select File --"] + file_names
        )

    remaining = [
        f for f in file_names
        if f != media_plan_name
    ]

    with col2:
        craft_name = st.selectbox(
            "📌 Craft Report",
            ["-- Select File --"] + remaining
        )

    used = [media_plan_name, craft_name]

    publisher_options = [
        f for f in file_names
        if f not in used
    ]

    with col3:
        publisher_names = st.multiselect(
            "📌 Publisher Reports",
            publisher_options
        )

    st.markdown("---")

    # =====================================================
    # STEP 1 BUTTON
    # =====================================================
    if "step1_done" not in st.session_state:
        st.session_state.step1_done = False

    if st.button("🚀 Step 1: Clean Publisher Files"):
        st.session_state.step1_done = True

    if not st.session_state.step1_done:
        st.stop()

    # =====================================================
    # VALIDATION
    # =====================================================
    if media_plan_name == "-- Select File --":
        st.error("Select Media Plan")
        st.stop()

    if craft_name == "-- Select File --":
        st.error("Select Craft Report")
        st.stop()

    if len(publisher_names) == 0:
        st.error("Select Publisher Reports")
        st.stop()

    # =====================================================
    # STORE FILES
    # =====================================================
# =====================================================
# STORE FILES
# =====================================================
    media_file = next(
        f for f in uploaded_files
        if f.name == media_plan_name
    )

    craft_file = next(
        f for f in uploaded_files
        if f.name == craft_name
    )

    publisher_files = [
        f for f in uploaded_files
        if f.name in publisher_names
    ]

    # SAVE TO SESSION STATE
    st.session_state.media_file = media_file
    st.session_state.craft_file = craft_file
    st.session_state.publisher_files = publisher_files
    # =====================================================
    # HELPERS
    # =====================================================
    def normalize(txt):
        return re.sub(r"[^a-z0-9]", "", str(txt).lower()).strip()

    # =====================================================
    # FORMAT CONFIG
    # =====================================================
    FORMATS = {

        "Google": {
            "identify": [
                "day","campaign","impr","clicks","views","leads"
            ],
            "mapping": {
                "Unique Key": ["campaign"],
                "Date": ["day"],
                "Impressions": ["impr"],
                "Clicks": ["clicks"],
                "Views": ["views"],
                "Spends": ["cost","spend"],
                "Engagements": ["postengagements"]
            }
        },

        "Sizmek": {
            "identify": [
                "day","placementname",
                "impressionsnet","clicksnet"
            ],
            "mapping": {
                "Unique Key": ["placementname"],
                "Date": ["day"],
                "Impressions": ["impressionsnet"],
                "Clicks": ["clicksnet"],
                "Views": ["views"],
                "Spends": ["cost","spend"],
                "Engagements": ["postengagements"]
            }
        },

        "Ecom": {
            "identify": [
                "activity","portfolio","category",
                "series","models","targeting",
                "month","ro","date",
                "campaign","type","views",
                "clicks","uniquekey"
            ],
            "mapping": {
                "Unique Key": ["uniquekey"],
                "Date": ["date"],
                "Impressions": ["impression"],
                "Clicks": ["clicks"],
                "Views": ["views"],
                "Spends": ["spend"],
                "Engagements": ["postengagements"]
            }
        },

        "Social": {
            "identify": [
                "uniquekey","campaigntype",
                "buytpe","campaignname",
                "adsetname","finaldaydate",
                "impressions","clicksall",
                "thruplays"
            ],
            "mapping": {
                "Unique Key": ["uniquekey"],
                "Date": ["finaldaydate"],
                "Impressions": ["impressions"],
                "Clicks": ["clicksall","linkclicks"],
                "Views": ["thruplays"],
                "Spends": ["amountspent","spend"],
                "Engagements": ["postengagements"]
            }
        },

        "DV360": {
            "identify": [
                "date","insertionorder",
                "lineitem","advertiser",
                "impressions","clicks",
                "trueviewviews"
            ],
            "mapping": {
                "Unique Key": ["insertionorder"],
                "Date": ["date"],
                "Impressions": ["impressions"],
                "Clicks": ["clicks"],
                "Views": ["trueviewviews","completeviewsvideo"],
                "Spends": ["cost","mediacost"],
                "Engagements": ["postengagements"]
            }
        },

        "Ecom-Display": {
            "identify": [
                "date",
                "advertiser",
                "order",
                "lineitem",
                "impressions",
                "clickthroughs",
                "qt",
                "campaignname",
                "unique",
                "views",
                "clicks"
            ],
            "mapping": {
                "Unique Key": ["unique","lineitemcomments"],
                "Date": ["date"],
                "Impressions": ["impressions"],
                "Clicks": ["clickthroughs","clicks","sum(clicks)"],
                "Views": ["views","sum(views)"],
                "Spends": ["spend","cost"],
                "Engagements": ["engagement"]
            }
        } ,
        "Ecom-Display2": {
            "identify": ["date","campaignname","qt","unique","sum(views)","sum(clicks)","ctr"],
            "mapping": {
                "Unique Key": ["unique"],
                "Date": ["date"],
                "Impressions": ["impressions"],
                "Clicks": ["clickthroughs","clicks","sum(clicks)"],
                "Views": ["views","sum(views)"],
                "Spends": ["spend","cost"],
                "Engagements": ["engagement"]
            }
        }   

    }

    # =====================================================
    # HELPERS
    # =====================================================
    def normalize(txt):
        return re.sub(r"[^a-z0-9]", "", str(txt).lower()).strip()


    # =====================================================
    # HEADER DETECTION (OLD STRONG LOGIC)
    # =====================================================
    def detect_header_row(raw):

        best_row = None
        best_score = 0

        for i in range(min(40, len(raw))):

            vals = raw.iloc[i].fillna("").astype(str).tolist()

            score = 0

            for val in vals:

                n = normalize(val)

                for k in [
                    "campaign", "day", "date", "impr",
                    "click", "view", "placement",
                    "unique", "lineitem", "advertiser"
                ]:

                    if n == k:
                        score += 3
                    elif k in n:
                        score += 1

            if score > best_score:
                best_score = score
                best_row = i

        return best_row


    # =====================================================
    # FORMAT DETECTION (OLD SMART LOGIC)
    # =====================================================
    def detect_format(df):

        cols = [normalize(c) for c in df.columns]

        best_format = None
        best_score = 0

        for fmt_name, cfg in FORMATS.items():

            score = 0
            total = len(cfg["identify"])

            for need in cfg["identify"]:

                if need in cols:
                    score += 3

                elif any(need in c for c in cols):
                    score += 1

            final_score = score / (total * 3)

            if final_score > best_score:
                best_score = final_score
                best_format = fmt_name

        if best_score >= 0.55:
            return best_format

        return None


    # =====================================================
    # COLUMN FINDER (OLD STRONG LOGIC)
    # =====================================================
    def get_col(df, options):

        cols = list(df.columns)

        # exact match first
        for opt in options:
            for c in cols:
                if normalize(c) == normalize(opt):
                    return c

        # partial match
        for opt in options:
            for c in cols:
                if normalize(opt) in normalize(c):
                    return c

        return None


    # =====================================================
    # CLEAN BY FORMAT
    # =====================================================
    def clean_by_format(df, fmt, file_name, sheet):

        out = pd.DataFrame()

        for final_col, options in FORMATS[fmt]["mapping"].items():

            src = get_col(df, options)

            if src:
                out[final_col] = df[src]
            else:
                out[final_col] = None

        out["Source File"] = file_name
        out["Sheet"] = sheet
        out["Format"] = fmt

        return out.reset_index(drop=True)


    # =====================================================
    # EXCEL PROCESSOR (OLD SAFE LOGIC)
    # =====================================================
    def process_excel(file):

        frames = []

        xl = pd.ExcelFile(file)

        for sheet in xl.sheet_names:

            try:
                raw = xl.parse(sheet, header=None)

                header_row = detect_header_row(raw)

                if header_row is None:
                    continue

                df = xl.parse(sheet, header=header_row)

                df = df.dropna(axis=1, how="all")
                df = df.dropna(how="all")

                # remove duplicate columns
                df = df.loc[:, ~df.columns.duplicated()]

                if len(df.columns) < 2:
                    continue

                fmt = detect_format(df)

                if fmt:

                    clean = clean_by_format(
                        df,
                        fmt,
                        file.name,
                        sheet
                    )

                    frames.append(clean)

            except:
                pass

        return frames


    # =====================================================
    # CSV PROCESSOR
    # =====================================================
    def process_csv(file):

        try:
            df = pd.read_csv(file)

            df = df.loc[:, ~df.columns.duplicated()]

            fmt = detect_format(df)

            if fmt:
                return [
                    clean_by_format(
                        df,
                        fmt,
                        file.name,
                        "CSV"
                    )
                ]

        except:
            pass

        return []
    # =====================================================
    # CLEANING START
    # =====================================================
# =====================================================
# CLEANING START
# =====================================================
    st.subheader("🧹 Cleaning Publisher Files")

    if "final_df" not in st.session_state:

        all_frames = []

        progress = st.progress(0)

        total = len(publisher_files)

        for i, file in enumerate(publisher_files):

            if file.name.lower().endswith(".csv"):
                frames = process_csv(file)
            else:
                frames = process_excel(file)

            all_frames.extend(frames)

            progress.progress((i + 1) / total)

        if not all_frames:
            st.error("No usable publisher data found.")
            st.stop()

        final_df = pd.concat(
            all_frames,
            ignore_index=True,
            sort=False
        )

        final_df = pd.concat(
            all_frames,
            ignore_index=True,
            sort=False
        )

        # ==================================
        # DATE FORMAT SAME AS PHASE 6
        # ==================================
        final_df["Date"] = pd.to_datetime(
            final_df["Date"],
            format="%Y/%m/%d",
            errors="coerce"
        ).fillna(
            pd.to_datetime(
                final_df["Date"],
                errors="coerce",
                dayfirst=True
            )
        )

        final_df = final_df.dropna(
            how="all",
            subset=[
                "Unique Key",
                "Date",
                "Impressions",
                "Clicks",
                "Views",
                "Engagements"
            ]
        )

        # SAVE CLEANED DATA
        st.session_state.final_df = final_df

    else:

        # USE SAVED DATA
        final_df = st.session_state.final_df
    # =====================================================
    # PREVIEW
    # =====================================================
    # st.success("✅ Publisher Cleaning Completed")

    # st.dataframe(
    #     final_df,
    #     use_container_width=True
    # )

    # =====================================================
    # DOWNLOAD CLEANED FILE
    # =====================================================
    excel_output = BytesIO()

    with pd.ExcelWriter(
        excel_output,
        engine="openpyxl"
    ) as writer:

        final_df.to_excel(
            writer,
            sheet_name="Publisher Data",
            index=False
        )

    st.download_button(
        "⬇️ Download Cleaned Publisher Data",
        data=excel_output.getvalue(),
        file_name="Unified_Publisher_Data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if "media_file" not in st.session_state:
        st.error("Please run Step 1 first.")
        st.stop()

    final_df = st.session_state.final_df
    media_file = st.session_state.media_file
    craft_file = st.session_state.craft_file
    publisher_files = st.session_state.publisher_files
    # =====================================================
    # STEP 2 BUTTON
    # =====================================================
    st.markdown("---")

    if st.button("🚀 Step 2: Generate Final SA Report"):
        st.session_state.step2_done = True

    if "step2_done" not in st.session_state:
        st.stop()

    if not st.session_state.step2_done:
        st.stop()

    # =====================================================
    # NOW CONTINUE PHASE 2 ONWARDS
    # =====================================================
    # media_file
    # craft_file
    # final_df
    # use final_df in Phase 6

    # return media_file, craft_file, final_df



    # Phase 2

    # =====================================================
    # PHASE 2 — MEDIA PLAN TAB EXTRACTION
    # =====================================================

    # media_file = files_by_type["media_plan"][0]
    media_xl = pd.ExcelFile(media_file)

    def normalize(x):
        return re.sub(r"[^a-z0-9]", "", str(x).lower())


    def is_data_sheet(name):
        n = normalize(name)
        return "checklist" not in n and "summary" not in n


    # ad_type_sheets = [s for s in media_xl.sheet_names if is_data_sheet(s)]
    ad_type_sheets = [
        s for s in media_xl.sheet_names
        if is_data_sheet(s) and normalize(s) != "estoresearch"
    ]

    st.subheader("📊 Media Plan Ad Types Detected")

    for s in ad_type_sheets:
        st.write("•", s)

    st.success(f"✅ {len(ad_type_sheets)} Ad Type tabs ready for processing")




    # Detect Brand name
    def detect_brand(filename):
        name = normalize(filename)

        if "watch" in name:
            return "Watch"
        if "mobile" in name or "phone" in name:
            return "Mobile"
        if "tv" in name:
            return "TV"

        return "Samsung"
        

    brand_name = detect_brand(media_file.name)
    def extract_qt(filename):
        match = re.search(r"qt\d+", filename.lower())
        if match:
            return match.group(0).upper()
        return "QTXXXXXXX"
    qt_number = extract_qt(media_file.name)




    campaign_start = None
    campaign_end = None

    for sheet in media_xl.sheet_names:

        if "checklist" in normalize(sheet):

            checklist = media_xl.parse(sheet, header=None)

            for _, row in checklist.iterrows():

                row_text = " ".join([str(x).lower() for x in row.values if pd.notna(x)])

                # ✅ ALWAYS compute dates (not inside condition)
                dates = pd.to_datetime(
                    pd.Series(row.values),
                    errors="coerce",
                    dayfirst=True
                ).dropna()

                if not dates.empty:

                    if "from" in row_text and campaign_start is None:
                        campaign_start = dates.iloc[0]

                    if "till" in row_text and campaign_end is None:
                        campaign_end = dates.iloc[0]

    if not campaign_start or not campaign_end:
        st.error("Could not detect campaign start/end dates from Checklist")
        st.stop()


    st.subheader("📆 Campaign Duration Detected")

    st.write("Start Date:", campaign_start.strftime("%d-%b-%Y"))
    st.write("End Date:", campaign_end.strftime("%d-%b-%Y"))

    campaign_days = (campaign_end - campaign_start).days + 1

    st.success(f"✅ Campaign Days: {campaign_days}")








    # =====================================================
    # PHASE 4 — MEDIA PLAN LINE ITEM PARSER (FIXED)
    # =====================================================

    def find_column_strict(df, must_have, optional=None):
        """
        must_have: list of mandatory keywords (eg ["click"])
        optional: list of optional keywords (eg ["est", "planned"])
        """
        for col in df.columns:
            col_norm = normalize(col)

            if all(k in col_norm for k in must_have):
                if optional:
                    if any(o in col_norm for o in optional):
                        return col
                else:
                    return col
        return None


    media_plan_items = {}

    for sheet in ad_type_sheets:

        raw = media_xl.parse(sheet, header=None)

        # ---- detect header row ----
        header_row = None
        for i in range(min(25, len(raw))):
            row = raw.iloc[i]

            row_text = " ".join([
                str(x).lower() for x in row.values if pd.notna(x)
            ])
            
            if "unique" in row_text and "buy" in row_text:
                header_row = i
                break

        if header_row is None:
            continue

        df = media_xl.parse(sheet, header=header_row).dropna(how="all")

        # ---- column mapping (STRICT) ----
        col_map = {
            "publisher": find_column_strict(df, ["publisher"]) or find_column_strict(df, ["site"]),
            "unique_key": find_column_strict(df, ["unique", "key"]),
            "objective": find_column_strict(df, ["objective"]),
            "property": find_column_strict(df, ["property"]) or find_column_strict(df, ["placement"]),
            "ad_unit": find_column_strict(df, ["ad", "unit"]),
            "buy_type": find_column_strict(df, ["buy"]),

            # Planned metrics (STRICT)
            "est_clicks": find_column_strict(df, ["click"], ["est", "planned"]),
            # "est_imps": find_column_strict(df, ["impression"], ["est", "planned"]),
            "est_imps": (find_column_strict(df, ["impression"], ["est", "planned"])or find_column_strict(df, ["impression"])or find_column_strict(df, ["impressions"])),
            "est_views": find_column_strict(df, ["view"], ["est", "planned"]),
            "est_eng": find_column_strict(df, ["engagement"], ["est", "planned"]),
            "est_leads": find_column_strict(df, ["lead"], ["est", "planned"]),
        }

        rows = []

        for _, r in df.iterrows():

            # uk = r.get(col_map["unique_key"])
            # if pd.isna(uk):
            #     continue
            
            uk = str(r.get(col_map["unique_key"])).strip()
            site = str(r.get(col_map["publisher"])).strip()

            if uk == "" or uk.lower() == "nan":
                continue

            if "total" in uk.lower():
                continue



            buy = str(r.get(col_map["buy_type"])).lower()
            planned_v1 = None

            # ---- CORRECT BUY TYPE → METRIC MAPPING ----
            if "cpc" in buy:
                planned_v1 = r.get(col_map["est_clicks"])
            elif any(x in buy for x in ["cpm", "fixed", "flat", "cph"]):
                planned_v1 = r.get(col_map["est_imps"])

            elif "cpv" in buy:
                planned_v1 = r.get(col_map["est_views"])
            elif "cpe" in buy:
                planned_v1 = r.get(col_map["est_eng"])
            elif "cpl" in buy:
                planned_v1 = r.get(col_map["est_leads"])

            rows.append({
                "Genres": sheet,
                "Site": r.get(col_map["publisher"]),
                "Unique Key": uk,
                "Objective/Targeting": r.get(col_map["objective"]),
                "Property/Inventory": r.get(col_map["property"], "-"),
                "Ad Unit": r.get(col_map["ad_unit"]),
                "Cost Format": buy.upper(),
                "Planned Delivery v1": planned_v1,
            })

        media_plan_items[sheet] = pd.DataFrame(rows)


    st.success("✅ Phase 4 FIXED — Planned Delivery v1 now matches Buy Type correctly")













    # Phase 5

    # =====================================================
    # PHASE 5 — CRAFT REPORT INTEGRATION
    # =====================================================

    # craft_file = files_by_type["craft"][0]
    craft_raw = pd.read_excel(craft_file, header=None)

    # --- detect header row ---
    craft_header = None
    for i in range(min(25, len(craft_raw))):
        row = craft_raw.iloc[i].astype(str).str.lower()
        # if "unique" in " ".join(row) and "planned" in " ".join(row):
        row_text = " ".join(row.astype(str).str.lower())

        if "unique" in row_text and "planned" in row_text:
            craft_header = i
            break

    if craft_header is None:
        st.error("CRAFT header not detected")
        st.stop()

    craft_df = pd.read_excel(craft_file, header=craft_header)
    craft_df.columns = [str(c) for c in craft_df.columns]


    def find_craft_col(keyword):
        for col in craft_df.columns:
            if keyword in normalize(col):
                return col
        return None

    def find_reported(keyword):
        for col in craft_df.columns:
            name = normalize(col)
            if "reported" in name and keyword in name:
                return col
        return None


    # craft_cols = {
    #     "qt": find_craft_col("qt"),
    #     "channel": find_craft_col("channel"),
    #     "uk": find_craft_col("unique"),
    #     "p_click": find_craft_col("plannedclick"),
    #     "p_imp": find_craft_col("plannedimp"),
    #     "p_view": find_craft_col("plannedvideo"),
    #     "p_eng": find_craft_col("plannedeng"),
    #     "r_click": find_craft_col("click"),
    #     "r_imp": find_craft_col("imp"),
    #     "r_view": find_craft_col("video"),
    #     "r_eng": find_craft_col("eng"),
    # }


    def find_reported_only(keyword):
        for col in craft_df.columns:
            name = normalize(col)

            if keyword in name and "planned" not in name:
                return col
        return None



    craft_cols = {
        "qt": find_craft_col("qt"),
        "channel": find_craft_col("channel"),
        "uk": find_craft_col("unique"),

        # PLANNED (same as before)
        "p_click": find_craft_col("plannedclick"),
        "p_imp": find_craft_col("plannedimp"),
        "p_view": find_craft_col("plannedvideo"),
        "p_eng": find_craft_col("plannedeng"),

        # ✅ REPORTED — NON PLANNED ONLY
        "r_click": find_reported_only("click"),
        "r_imp": find_reported_only("impression"),
        "r_view": find_reported_only("video"),
        "r_eng": find_reported_only("engagement"),
    }






    def get_craft_values(ad_type, uk, buy_type):

        df = craft_df.copy()

        # ---------- TRY NORMAL FILTER (CHANNEL + UNIQUE) ----------
        if craft_cols["channel"]:
            df_filtered = df[
                df[craft_cols["channel"]]
                .astype(str)
                .str.contains(ad_type, case=False, na=False)
            ]
        else:
            df_filtered = df

        df_filtered = df_filtered[
            df_filtered[craft_cols["uk"]].astype(str) == str(uk)
        ]

        # ---------- FALLBACK: UNIQUE KEY ONLY ----------
        if df_filtered.empty:
            df_filtered = df[
                df[craft_cols["uk"]].astype(str) == str(uk)
            ]

        if df_filtered.empty:
            return None, None

        buy = buy_type.lower()

        if "cpc" in buy:
            return (
                df_filtered[craft_cols["p_click"]].sum(),
                df_filtered[craft_cols["r_click"]].sum()
            )

        if "cpm" in buy or "fixed" in buy or "flat" in buy or "cph" in buy:
            return (
                df_filtered[craft_cols["p_imp"]].sum(),
                df_filtered[craft_cols["r_imp"]].sum()
            )

        if "cpv" in buy:
            return (
                df_filtered[craft_cols["p_view"]].sum(),
                df_filtered[craft_cols["r_view"]].sum()
            )

        if "cpe" in buy:
            return (
                df_filtered[craft_cols["p_eng"]].sum(),
                df_filtered[craft_cols["r_eng"]].sum()
            )

        return None, None



    # =====================================================
    # ATTACH CRAFT VALUES TO MEDIA PLAN STRUCTURE
    # =====================================================

    for ad_type, df in media_plan_items.items():

        craft_planned = []
        craft_reported = []

        for _, row in df.iterrows():

            cp, cr = get_craft_values(
                ad_type,
                row["Unique Key"],
                row["Cost Format"]
            )

            craft_planned.append(cp)
            craft_reported.append(cr)

        df["CRAFT Planned Delivery"] = craft_planned
        df["CRAFT Reported Delivery"] = craft_reported


    st.success("✅ CRAFT planned & reported values mapped to all ad types")







        # ================================================s=====
        # PHASE 6 — VERIFIED PUBLISHER EXTRACTION (FINAL WORKING)
        # =====================================================
    # =====================================================
    # PHASE 6 — USE CLEANED DATA ONLY (UPDATED & CORRECTED)
    # =====================================================

    st.subheader("📊 Phase 6 — Pulling Data from Cleaned Publisher DF")


    def metric_from_buy_type(buy):

        b = str(buy).lower()

        if "cpc" in b:
            return "Clicks"

        elif "cpv" in b:
            return "Views"

        elif "cpe" in b:
            return "Engagements"

        elif "cpl" in b:
            return "Leads"

        else:
            return "Impressions"


    for ad_type, df in media_plan_items.items():

        actuals = []
        lives = []

        for _, row in df.iterrows():

            uk = str(row["Unique Key"]).strip()
            buy = row["Cost Format"]

            # =====================================
            # FILTER CLEANED FILE — EXACT MATCH FIRST
            # =====================================
            pub = final_df[
                final_df["Unique Key"]
                .astype(str)
                .apply(lambda x: normalize(x) == normalize(uk))
            ].copy()

            # fallback partial match if no exact rows
            if pub.empty:
                pub = final_df[
                    final_df["Unique Key"]
                    .astype(str)
                    .apply(lambda x: normalize(uk) in normalize(x))
                ].copy()

            if pub.empty:
                actuals.append(None)
                lives.append(None)
                continue

            # =====================================
            # METRIC COLUMN BASED ON BUY TYPE
            # =====================================
            metric_col = metric_from_buy_type(buy)

            if metric_col not in pub.columns:
                actuals.append(None)
                lives.append(None)
                continue

            # =====================================
            # ACTUAL VALUE
            # =====================================
            val = pd.to_numeric(
                pub[metric_col]
                .astype(str)
                .str.replace(",", "", regex=False)
                .str.replace("--", "", regex=False)
                .str.strip(),
                errors="coerce"
            ).sum()

            # =====================================
            # LIVE DATE = EARLIEST VALID DATE
            # =====================================
            live = pd.to_datetime(
                pub["Date"],
                errors="coerce",
                dayfirst=True
            ).dropna().min()

            actuals.append(val if val > 0 else None)
            lives.append(live if pd.notna(live) else None)

        df["Actual Delivered Reporting SA"] = actuals
        df["Live Date"] = lives

    st.success("✅ Phase 6 COMPLETE — Correct Data Pulled from Cleaned DF")


    # phase 7
    # =====================================================
    # PHASE 7 — SAFE SA CALCULATIONS (AUTO COLUMN DETECTION)
    # =====================================================




    def normalize(x):
        return re.sub(r"[^a-z0-9]", "", str(x).lower())


    def find_col(df, keyword):
        for c in df.columns:
            if keyword in normalize(c):
                return c
        return None


    def safe_div(a, b):
        if pd.isna(a) or pd.isna(b) or b == 0:
            return None
        return a / b


    def safe_diff_div(a, b, base):
        if pd.isna(a) or pd.isna(b) or pd.isna(base) or base == 0:
            return None
        return (a - b) / base


    for ad_type, df in media_plan_items.items():

        col_actual = find_col(df, "actualdelivered")
        col_planned = find_col(df, "planneddeliveryv1")
        col_craft_plan = find_col(df, "craftplanneddelivery")
        col_craft_rep = find_col(df, "craftreporteddelivery")

        if not all([col_actual, col_planned, col_craft_plan, col_craft_rep]):
            st.warning(f"⚠ Missing columns in {ad_type} — skipping calculations")
            continue

        actual = df[col_actual]
        planned_v1 = df[col_planned]
        craft_planned = df[col_craft_plan]
        craft_reported = df[col_craft_rep]

        df["% v1 Delivery"] = [
            safe_div(a, p) for a, p in zip(actual, planned_v1)
        ]

        df["% Final Delivery"] = [
            safe_div(a, cp) for a, cp in zip(actual, craft_planned)
        ]

        df["Total KPI Achieved"] = df["% Final Delivery"]

        df["Deviation % v1 & CRAFT Plan"] = [
            safe_diff_div(cp, p, p)
            for cp, p in zip(craft_planned, planned_v1)
        ]

        df["Deviation % Platform & CRAFT Delivery"] = [
            safe_diff_div(a, cr, cr)
            for a, cr in zip(actual, craft_reported)
        ]


    st.success("✅ Phase 7 completed — calculations added safely for all tabs")












    # =====================================================
    # PHASE 8 — FINAL EXCEL OUTPUT + FORMATTING
    # =====================================================



    FINAL_COLUMNS = [
        "Genres","Site","Unique Key","Objective/Targeting","Property/Inventory",
        "Ad Unit","Cost Format","Campaign Days","Monitoring Days",
        "Start Date","End Date","Live Date",
        "Planned Delivery v1","CRAFT Planned Delivery",
        "Actual Delivered Reporting SA","CRAFT Reported Delivery",
        "% v1 Delivery","% Final Delivery","Total KPI Achieved",
        "Deviation % v1 & CRAFT Plan","Deviation % Platform & CRAFT Delivery",
    ]


    # ---------- Add campaign & monitoring days ----------

    for df in media_plan_items.values():
        

        df["Start Date"] = campaign_start
        df["End Date"] = campaign_end

        df["Campaign Days"] = (campaign_end - campaign_start).days + 1
        # df["Monitoring Days"] = (campaign_end - campaign_start).days + 1

        df["Monitoring Days"] = df["Live Date"].apply(
            lambda x: (campaign_end - x).days + 1 if pd.notna(x) else None
        )


    # ---------- Write Excel ----------

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        master_frames = []

        for ad_type, df in media_plan_items.items():

            df_final = df.reindex(columns=FINAL_COLUMNS)

            # df_final.to_excel(writer, sheet_name=ad_type[:31], index=False)
            sheet_name = ad_type[:31]

            df_final.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=7   # pushes table down
            )

            ws = writer.sheets[sheet_name]


            # ---- Top info block ----

            
            ws["A1"] = "Client"
            ws["B1"] = "Samsung India"

            ws["A2"] = "Brand"
            ws["B2"] = brand_name

            ws["A3"] = "Month"
            ws["B3"] = campaign_start.strftime("%b %Y")

            ws["A4"] = "Monitoring Period"
            ws["B4"] = f"{campaign_start.strftime('%d %b')} – {campaign_end.strftime('%d %b')}"

            ws["A5"] = "Campaign Duration"
            ws["B5"] = f"{campaign_start.strftime('%d %b')} – {campaign_end.strftime('%d %b')}"


            

            HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
            BOLD = Font(bold=True)

            for row in range(1, 6):
                ws[f"A{row}"].font = BOLD
                ws[f"A{row}"].fill = HEADER_FILL
                ws[f"B{row}"].fill = HEADER_FILL

            master_frames.append(df_final)

        if master_frames:
            pd.concat(master_frames).to_excel(
                writer, sheet_name="Master", index=False
            )


    # ---------- Formatting ----------
    output.seek(0)
    wb = load_workbook(output, data_only=True)

    

    DARK_BLUE = PatternFill("solid", fgColor="FF203764")
    LIGHT_BLUE = PatternFill("solid", fgColor="FF4472C4")
    GREY = PatternFill("solid", fgColor="FF808080")
    GREEN = PatternFill("solid", fgColor="FF70AD47")
    LIGHT_GREY = PatternFill("solid", fgColor="FF808080")

    HEADER_FONT = Font(size=12, color="FFFFFF", bold=True)
    BODY_FONT = Font(size=12)
    # bold=True,
    DETAIL_LABEL = Font(size=10, color="FFFFFF")
    DETAIL_VALUE = Font(size=10, color="FFFFFF")

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


    BLUE_COLS = [
    "Genres","Site","Unique Key","Objective/Targeting","Property/Inventory",
    "Ad Unit","Cost Format","Campaign Days","Monitoring Days",
    "Start Date","End Date","Live Date"
    ]

    LIGHT_BLUE_COL = ["Deviation % v1 & CRAFT Plan"]

    GREY_COLS = [
    "Planned Delivery v1",
    "Actual Delivered Reporting SA",
    "% v1 Delivery",
    "% Final Delivery",
    "Total KPI Achieved"
    ]

    GREEN_COLS = [
    "CRAFT Planned Delivery",
    "CRAFT Reported Delivery",
    "Deviation % Platform & CRAFT Delivery"
    ]


    for ws in wb.worksheets:

        # ===== Header details section =====
        if ws.title != "Master":
            for r in range(1,6):
                ws[f"A{r}"].font = DETAIL_LABEL
                ws[f"B{r}"].font = DETAIL_VALUE
                ws[f"A{r}"].fill = LIGHT_GREY
                ws[f"B{r}"].fill = LIGHT_GREY
                ws[f"A{r}"].border = BORDER
                ws[f"B{r}"].border = BORDER

        header_row = 1 if ws.title == "Master" else 8
        data_row = header_row + 1

        headers = {
            ws.cell(header_row, c).value: c
            for c in range(1, ws.max_column + 1)
        }

        # ===== Format header row =====
        for name, col in headers.items():

            cell = ws.cell(header_row, col)

            if name in BLUE_COLS:
                cell.fill = DARK_BLUE

            elif name in LIGHT_BLUE_COL:
                cell.fill = LIGHT_BLUE

            elif name in GREY_COLS:
                cell.fill = GREY

            elif name in GREEN_COLS:
                cell.fill = GREEN

            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

        # ===== Table body =====
        for r in range(data_row, ws.max_row + 1):
            # for c in range(1, ws.max_column + 1):
            last_col = len(headers)
            for c in range(1, last_col + 1):        
                cell = ws.cell(r, c)
                cell.font = BODY_FONT
                # cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.alignment = CENTER
                cell.border = BORDER
                        # 👉 Apply Indian number format
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##,##0'

    FILL_UNDER = PatternFill("solid", fgColor="FFC000")
    FILL_OVER = PatternFill("solid", fgColor="FFCCFF")
    FILL_NA = PatternFill("solid", fgColor="9933FF")

    FONT_GREEN = Font(color="00B050")
    FONT_RED = Font(color="FF0000")


    def get_threshold(sheet):
        return 0.30 if "dmp" in sheet.lower() else 0.10

    for ws in wb.worksheets:

        # ---------- MASTER SHEET ----------
        if ws.title == "Master":

            TABLE_HEADER_ROW = 1
            DATA_START_ROW = 2

        # ---------- AD TABS ----------
        else:

            TABLE_HEADER_ROW = 8
            DATA_START_ROW = TABLE_HEADER_ROW + 1

        headers = {cell.value: i+1 for i, cell in enumerate(ws[TABLE_HEADER_ROW])}
        WHOLE_PERCENT = [
            "% v1 Delivery",
            "% Final Delivery",
            "Total KPI Achieved",
            "Deviation % v1 & CRAFT Plan"
        ]

        TWO_DECIMAL_PERCENT = [
            "Deviation % Platform & CRAFT Delivery"
        ]

        # -------- Whole % columns --------
        for col in WHOLE_PERCENT:
            if col in headers:
                col_index = headers[col]
                # for r in range(2, ws.max_row + 1):
                DATA_START_ROW = TABLE_HEADER_ROW + 1

                for r in range(DATA_START_ROW, ws.max_row + 1):
                    cell = ws.cell(r, col_index)
                    if cell.value is not None:
                        cell.value = float(cell.value)
                        cell.number_format = "0%"

        # -------- 2 Decimal % column --------
        for col in TWO_DECIMAL_PERCENT:
            if col in headers:
                col_index = headers[col]
                # for r in range(2, ws.max_row + 1):
                DATA_START_ROW = TABLE_HEADER_ROW + 1

                for r in range(DATA_START_ROW, ws.max_row + 1):
                    cell = ws.cell(r, col_index)
                    if cell.value is not None:
                        cell.value = float(cell.value)
                        cell.number_format = "0.00%"


        col_final = headers["% Final Delivery"]
        col_kpi = headers["Total KPI Achieved"]
        col_dev = headers["Deviation % Platform & CRAFT Delivery"]

        threshold = get_threshold(ws.title)

        # for r in range(2, ws.max_row + 1):
        DATA_START_ROW = TABLE_HEADER_ROW + 1

        for r in range(DATA_START_ROW, ws.max_row + 1):

            val = ws.cell(r, col_final).value

            if val is None:
                ws.cell(r, col_final).fill = FILL_NA
            elif val < (1 - threshold):
                ws.cell(r, col_final).fill = FILL_UNDER
            elif val > (1 + threshold):
                ws.cell(r, col_final).fill = FILL_OVER

            # ws.cell(
            #     r, col_kpi
            # ).font = FONT_GREEN if val and val >= (1 - threshold) else FONT_RED

            dev = ws.cell(r, col_dev).value
            if dev is not None and abs(dev) > 0.02:
                ws.cell(r, col_dev).font = FONT_RED

        # Date formatting
        # for col_name in ["Start Date", "End Date", "Live Date"]:
        #     c = headers[col_name]
        #     for r in range(DATA_START_ROW, ws.max_row + 1):
        #         ws.cell(r, c).number_format = "DD-MMM-YY"

        for col_name in ["Start Date", "End Date", "Live Date"]:
            c = headers[col_name]
            for r in range(DATA_START_ROW, ws.max_row + 1):
                ws.cell(r, c).number_format = "DD-MMM"
        
        

        # ================= TOTAL ROW (AD TABS ONLY) =================
        # TOTAL FOR ALL SHEETS (remove condition)
        last_data_row = ws.max_row

        # 🔥 If Master — ignore blank rows
        while last_data_row > DATA_START_ROW and ws.cell(last_data_row, 1).value in [None, ""]:
            last_data_row -= 1
        total_row = last_data_row + 1

        ws.cell(total_row, 1).value = "Total"
        ws.cell(total_row, 1).fill = DARK_BLUE
        ws.cell(total_row, 1).font = HEADER_FONT
        ws.cell(total_row, 1).alignment = CENTER
        ws.cell(total_row, 1).border = BORDER

        SUM_COLS = [
                "Planned Delivery v1",
                "CRAFT Planned Delivery",
                "Actual Delivered Reporting SA",
                "CRAFT Reported Delivery",
            ]

        for col_name in SUM_COLS:
                if col_name in headers:
                    col = headers[col_name]
                    letter = get_column_letter(col)

                    # ws.cell(
                    #     total_row, col
                    # ).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
                    cell = ws.cell(total_row, col)
                    cell.value = f"=SUM({letter}{DATA_START_ROW}:{letter}{last_data_row})"
                
                # 👉 APPLY INDIAN FORMAT
                    cell.number_format = '#,##,##0'
        # ===== % TOTALS AS AVERAGE (NOT SUM) =====

        # ================= TOTAL % & DEVIATIONS — FROM TOTAL NUMBERS =================

        col_plan = headers["Planned Delivery v1"]
        col_craft_plan = headers["CRAFT Planned Delivery"]
        col_actual = headers["Actual Delivered Reporting SA"]
        col_craft_rep = headers["CRAFT Reported Delivery"]

        col_pct_v1 = headers["% v1 Delivery"]
        col_pct_final = headers["% Final Delivery"]
        col_kpi = headers["Total KPI Achieved"]
        col_dev_plan = headers["Deviation % v1 & CRAFT Plan"]
        col_dev_platform = headers["Deviation % Platform & CRAFT Delivery"]

        L_plan = get_column_letter(col_plan)
        L_craft_plan = get_column_letter(col_craft_plan)
        L_actual = get_column_letter(col_actual)
        L_craft_rep = get_column_letter(col_craft_rep)

        # % v1 Delivery
        ws.cell(total_row, col_pct_v1).value = f"={L_actual}{total_row}/{L_plan}{total_row}"

        # % Final Delivery
        ws.cell(total_row, col_pct_final).value = f"={L_actual}{total_row}/{L_craft_plan}{total_row}"

        # KPI Achieved
        # ws.cell(total_row, col_kpi).value = f"={get_column_letter(col_pct_final)}{total_row}"
        # KPI Achieved (CORRECT LOGIC)
        ws.cell(total_row,col_kpi).value = f"=SUM({L_actual}{DATA_START_ROW}:{L_actual}{last_data_row})/SUM({L_craft_plan}{DATA_START_ROW}:{L_craft_plan}{last_data_row})"

        # Deviation v1 vs CRAFT plan
        ws.cell(
            total_row, col_dev_plan
        ).value = f"=({L_craft_plan}{total_row}-{L_plan}{total_row})/{L_plan}{total_row}"

        # Deviation Platform vs CRAFT reported
        ws.cell(
            total_row, col_dev_platform
        ).value = f"=({L_actual}{total_row}-{L_craft_rep}{total_row})/{L_craft_rep}{total_row}"

        # formatting
        ws.cell(total_row, col_pct_v1).number_format = "0%"
        ws.cell(total_row, col_pct_final).number_format = "0%"
        ws.cell(total_row, col_kpi).number_format = "0%"
        ws.cell(total_row, col_dev_plan).number_format = "0%"
        ws.cell(total_row, col_dev_platform).number_format = "0.00%"

            # ================= FIXED KPI MERGE (FINAL) =================

        if ws.title != "Master":
        
            col_kpi = headers.get("Total KPI Achieved")
            col_actual = headers.get("Actual Delivered Reporting SA")
            col_craft_plan = headers.get("CRAFT Planned Delivery")
        
            if col_kpi and col_actual and col_craft_plan:
        
                # ✅ STEP 1: Calculate KPI manually (NOT Excel formula)
                total_actual = 0
                total_plan = 0
        
                for r in range(DATA_START_ROW, last_data_row + 1):
                    a = ws.cell(r, col_actual).value
                    p = ws.cell(r, col_craft_plan).value
        
                    try:
                        if a:
                            total_actual += float(a)
                        if p:
                            total_plan += float(p)
                    except:
                        continue
        
                kpi_value = None
                if total_plan != 0:
                    kpi_value = total_actual / total_plan
        
                # ✅ STEP 2: CLEAR column BEFORE merge
                for r in range(DATA_START_ROW, last_data_row + 1):
                    ws.cell(r, col_kpi).value = None
        
                # ✅ STEP 3: MERGE
                ws.merge_cells(
                    start_row=DATA_START_ROW,
                    end_row=last_data_row,
                    start_column=col_kpi,
                    end_column=col_kpi
                )
        
                # ✅ STEP 4: SET VALUE
                cell = ws.cell(DATA_START_ROW, col_kpi)
                cell.value = kpi_value
        
                # ✅ STEP 5: FORMAT (NO BACKGROUND as you said)
                cell.number_format = "0%"
                    # APPLY COLOR BASED ON KPI
                if kpi_value is not None:
                
                    if kpi_value >= (1 - threshold):
                        font_color = "00B050"   # GREEN
                    else:
                        font_color = "FF0000"   # RED
                
                    cell.font = Font(size=12, color=font_color)
                
                else:
                    cell.font = Font(size=12)
                # cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")

       
        



            
        
            # ✅ FILL TOTAL ROW ONLY ACROSS TABLE (NOT FULL SHEET)
        
        last_table_col = ws.max_column   # this is your table width
        for c in range(1, last_table_col + 1):
                cell = ws.cell(total_row, c)
                cell.fill = DARK_BLUE
                cell.font = HEADER_FONT
                cell.alignment = CENTER
                cell.border = BORDER            

            
        
# ================= MERGE GENRES + DIVIDER =================

        last_col = len(headers)
        
        if "Genres" in headers:
        
            THICK = Side(style="medium")
        
            col_genre = headers["Genres"]
        
            start_row_merge = DATA_START_ROW
            prev_value = ws.cell(DATA_START_ROW, col_genre).value
        
            for r in range(DATA_START_ROW + 1, last_data_row + 1):
        
                current_value = ws.cell(r, col_genre).value
        
                # ===== NEW GENRE START =====
                if current_value != prev_value:
        
                    # ===== MERGE PREVIOUS BLOCK =====
                    if start_row_merge < r - 1:
        
                        ws.merge_cells(
                            start_row=start_row_merge,
                            end_row=r - 1,
                            start_column=col_genre,
                            end_column=col_genre
                        )
        
                        merged_cell = ws.cell(start_row_merge, col_genre)
        
                        merged_cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center"
                        )
        
                    # ===== THICK DIVIDER =====
                    for c in range(1, last_col + 1):
        
                        cell = ws.cell(r, c)
        
                        cell.border = Border(
                            top=THICK,
                            left=cell.border.left,
                            right=cell.border.right,
                            bottom=cell.border.bottom
                        )
        
                    start_row_merge = r
                    prev_value = current_value
        
            # ===== LAST MERGE =====
            if start_row_merge < last_data_row:
        
                ws.merge_cells(
                    start_row=start_row_merge,
                    end_row=last_data_row,
                    start_column=col_genre,
                    end_column=col_genre
                )
        
                merged_cell = ws.cell(start_row_merge, col_genre)
        
                merged_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
        
        
        # ================= PERFECT OUTER TABLE BORDER =================
        
        THICK_SIDE = Side(style="medium")
        
        # ===== TOP BORDER =====
        for c in range(1, last_col + 1):
        
            cell = ws.cell(TABLE_HEADER_ROW, c)
        
            cell.border = Border(
                top=THICK_SIDE,
                left=cell.border.left,
                right=cell.border.right,
                bottom=cell.border.bottom
            )
        
        # ===== BOTTOM BORDER =====
        for c in range(1, last_col + 1):
        
            cell = ws.cell(total_row, c)
        
            cell.border = Border(
                bottom=THICK_SIDE,
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top
            )
        
        # ===== LEFT BORDER =====
        for r in range(TABLE_HEADER_ROW, total_row + 1):
        
            cell = ws.cell(r, 1)
        
            cell.border = Border(
                left=THICK_SIDE,
                top=cell.border.top,
                right=cell.border.right,
                bottom=cell.border.bottom
            )
        
        # ===== RIGHT BORDER =====
        for r in range(TABLE_HEADER_ROW, total_row + 1):
        
            cell = ws.cell(r, last_col)
        
            cell.border = Border(
                right=THICK_SIDE,
                top=cell.border.top,
                left=cell.border.left,
                bottom=cell.border.bottom
            )



        
        if ws.title != "Master":
        
            col_final = headers["% Final Delivery"]
            col_uk = headers["Unique Key"]
        
            under = []
            over = []
            perfect = []
        
            threshold = get_threshold(ws.title)
        
            for r in range(DATA_START_ROW, last_data_row + 1):
        
                val = ws.cell(r, col_final).value
                uk = ws.cell(r, col_uk).value
        
                if val is None:
                    continue
        
                # convert safely to float
                try:
                    val = float(val)
                except:
                    continue
        
                # normalize like Excel display
                val_check = round(val, 2)
        
                # ✅ PERFECT DELIVERY
                if val_check == 1:
                    perfect.append(str(uk))
        
                # 🔻 UNDER
                elif val < (1 - threshold):
                    under.append(str(uk))
        
                # 🔺 OVER
                elif val > (1 + threshold):
                    over.append(str(uk))
        
            # ================= REMARKS =================
            remarks = ["SA Remarks:"]
        
            if perfect:
                remarks.append(
                    "KPI delivered for unique key" +
                    ("s " if len(perfect) > 1 else " ") +
                    ", ".join(perfect)
                )
        
            if under:
                remarks.append(
                    "KPI under delivered for unique key" +
                    ("s " if len(under) > 1 else " ") +
                    ", ".join(under)
                )
        
            if over:
                remarks.append(
                    "KPI over delivered for unique key" +
                    ("s " if len(over) > 1 else " ") +
                    " and ".join(over)
                )
        
            # ================= WRITE TO EXCEL =================
            remarks_row = total_row + 2
        
            ws.merge_cells(
                start_row=remarks_row,
                start_column=1,
                end_row=remarks_row,
                end_column=ws.max_column
            )
        
            ws.cell(remarks_row, 1).value = "\n".join(remarks)
        
            ws.cell(remarks_row, 1).font = Font(size=10, italic=True)
            ws.cell(remarks_row, 1).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

    final_output = BytesIO()
    wb.save(final_output)


    st.success("✅ SA Report Generated Successfully")

    st.download_button(
        "⬇️ Download SA Report",
        data=final_output.getvalue(),
        file_name=f"SA_Report_{brand_name}_{qt_number}.xlsx",
        # file_name="SA_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




    return final_output.getvalue(), brand_name, qt_number





if __name__ == "__main__":
    run_sa_report()
